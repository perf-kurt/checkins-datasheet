from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Function that defines columns based on header titles


def defineColumn(headerSheetName, headers, headerrow):
    for col in range(1, headerSheetName.max_column + 1):
        for header in headers:
            if headerSheetName[get_column_letter(col) + str(headerrow)].value != None:
                if headerSheetName[get_column_letter(col) + str(headerrow)].value.replace('\n', ' ').upper().lstrip().rstrip() == header.upper():
                    return get_column_letter(col)


# Workbooks
classes_wb = load_workbook(
    filename='utils/EnrollmentDetailRpt.xlsx')
classes = classes_wb.active

# Classes columns
teachers_col = defineColumn(classes, ['Instructors'], 1)
level_col = defineColumn(classes, ['Cat1'], 1)
drop_col = defineColumn(classes, ['Drop'], 1)
class_col = defineColumn(classes, ['Class Name'], 1)
first_name_col = defineColumn(classes, ['Student First Name'], 1)
last_name_col = defineColumn(classes, ['Student Last Name'], 1)

# Create list of teachers, students and levels
enrollments = []
teachers_list = []

print()
print('Reading class file from jackrabbit and builing the data input file...')
print()

for row in range(2, classes.max_row+1):

    # Skip blank rows
    if classes[class_col+str(row)].value != None:

        _level = classes[level_col+str(row)].value
        _teachers = classes[teachers_col+str(row)].value
        if _teachers != None:
            # Select only first teacher name in the list
            _teacher = _teachers.split(',')[0]
        _first_name = classes[first_name_col+str(row)].value
        _last_name = classes[last_name_col+str(row)].value
        _class = classes[class_col+str(row)].value

        if _level not in ['Adults', 'Company', 'misc'] and _teacher != None:

            _tier = ''

            if _level in ['MiniMovers', 'Newbies', 'Petites']:
                _tier = 'Move'

            if _level in ['Minis', 'Beginners']:
                _tier = 'Develop'

            if _level in ['Intermediate', 'Advanced', 'Intermediate/Advanced']:
                _tier = 'Connect'

            # Only add the enrollment if the name/teacher/level combination is unique
            if any(d['name'] == _last_name + ', ' + _first_name and d['teacher'] == _teacher and d['level'] == _level for d in enrollments):
                continue

            enrollments.append({
                "name": _last_name + ', ' + _first_name,
                "teacher": _teacher,
                "level": _level,
                "tier": _tier,
                "class": _class
            })

            if _teacher not in teachers_list:
                teachers_list.append(_teacher)

# Create check in card data sheet using teacher objects
template_wb = load_workbook(
    filename='utils/Datasheet Template.xlsx')
    
for teacher in teachers_list:

    if teacher == '':
        continue

    # Create sheets for each tier
    _move = template_wb.copy_worksheet(template_wb['Move'])
    _develop = template_wb.copy_worksheet(template_wb['Develop'])
    _connect = template_wb.copy_worksheet(template_wb['Connect'])

    _move.title = teacher + ' - Move'
    _develop.title = teacher + ' - Develop'
    _connect.title = teacher + ' - Connect'

    for student in filter(lambda x: x['teacher'] == teacher, enrollments):

        # Move
        if student['tier'] == 'Move':
            _move['A'+str(_move.max_row+1)] = student['name']
            _move['B'+str(_move.max_row)] = student['level']

        # Develop
        elif student['tier'] == 'Develop':
            _develop['A'+str(_develop.max_row+1)] = student['name']
            _develop['B'+str(_develop.max_row)] = student['level']

        # Connect
        elif student['tier'] == 'Connect':
            _connect['A'+str(_connect.max_row+1)] = student['name']
            _connect['B'+str(_connect.max_row)] = student['level']

    # Remove empty sheets
    if _move.max_row == 6:
        template_wb.remove(_move)
    if _develop.max_row == 6:
        template_wb.remove(_develop)
    if _connect.max_row == 6:
        template_wb.remove(_connect)
    print('Created check-in sheets for ' + teacher)

# Remove template sheets
template_wb.remove(template_wb['Move'])
template_wb.remove(template_wb['Develop'])
template_wb.remove(template_wb['Connect'])

# ###### ###### ##    ## ######
# ##     ##  ## ##    ## ##
# ###### ###### ##    ## #####
#     ## ##  ##  ##  ##  ##
# ###### ##  ##    ##    ######

template_wb.save('exports/' + str(datetime.now().year) + ' - ' +  str(datetime.now().year+1) + ' Report Cards - Data Entry.xlsx')

print('Check-ins workbook created for ' + teacher)

input('\nCheck workbooks created! Press enter to exit')


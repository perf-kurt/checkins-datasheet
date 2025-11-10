# Iterate files in /exports
import os
from openpyxl import load_workbook

# Load mail merge datasheet template
mail_merge_data_wb = load_workbook(filename='utils/Mail Merge Data Template.xlsx')
data_entry_wb = load_workbook(filename='imports/Data Entry.xlsx')
mm_sheet = mail_merge_data_wb.active

# Read data from each sheet in data entry file
cards = []

for sheet in data_entry_wb.worksheets:
    # Read data from each row in the sheet
    for row in sheet.iter_rows(min_row=7, values_only=True):
        if row[0] is not None:  # Make sure data exists

            card_data = {
                'File Name': sheet.title.split(' - ')[0],
                'First Name': row[0],
                'Last Name': row[1],
                'Address 1': row[2],
                'Address 2': row[3],
                'City': row[4],
                'State': row[5],
                'Zip': row[6],
                'Country': row[7],
                'Tier': sheet.title.split(' - ')[1],
                'Attr 1': sheet['C6'].value,
                'Attr 2': sheet['D6'].value,
                'Attr 3': sheet['E6'].value,
                'Attr 4': sheet['F6'].value,
                'Attr 5': sheet['G6'].value,
                'Class': sheet['H6'].value,
            }
            cards.append(card_data)

    print("Building list of cards using data from sheet:", sheet.title)

# Order cards by Last Name, First Name and assign card #
cards = sorted(cards, key=lambda x: (x['Last Name'], x['First Name']))
card_no = 1
for card in cards:
    card['Card'] = card_no
    card_no += 1

# Assign page numbers
page = 1
breakpoint = (len(cards) // 4) + 1
for card in cards:
    if card['Card'] % breakpoint == 0:
        page = 1
        card['Page'] = page
    else:
        card['Page'] = page
    page += 1

print("\n","Total cards to process: ", len(cards),"\n")
print("Populating mail merge datasheet...","\n")
for card_data in cards:           
    mm_sheet['A' + str(mm_sheet.max_row + 1)] = card_data['File Name']  # File name without extension
    mm_sheet['B' + str(mm_sheet.max_row)] = card_data['First Name']
    mm_sheet['C' + str(mm_sheet.max_row)] = card_data['Last Name']
    mm_sheet['D' + str(mm_sheet.max_row)] = card_data['Address 1']
    mm_sheet['E' + str(mm_sheet.max_row)] = card_data['Address 2']
    mm_sheet['F' + str(mm_sheet.max_row)] = card_data['City']
    mm_sheet['G' + str(mm_sheet.max_row)] = card_data['State']
    mm_sheet['H' + str(mm_sheet.max_row)] = card_data['Zip']
    mm_sheet['I' + str(mm_sheet.max_row)] = card_data['Country']
    mm_sheet['J' + str(mm_sheet.max_row)] = card_data['Tier']
    mm_sheet['K' + str(mm_sheet.max_row)] = card_data['Attr 1']
    mm_sheet['L' + str(mm_sheet.max_row)] = card_data['Attr 2']
    mm_sheet['M' + str(mm_sheet.max_row)] = card_data['Attr 3']
    mm_sheet['N' + str(mm_sheet.max_row)] = card_data['Attr 4']
    mm_sheet['O' + str(mm_sheet.max_row)] = card_data['Attr 5']
    mm_sheet['P' + str(mm_sheet.max_row)] = card_data['Card']
    mm_sheet['Q' + str(mm_sheet.max_row)] = card_data['Page']


mail_merge_data_wb.save('Mail Merge Datasheet.xlsx')